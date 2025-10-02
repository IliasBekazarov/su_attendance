from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, authenticate, logout as auth_logout
from django.contrib.auth.forms import UserCreationForm
from django.db.models import Q, Count
from django.contrib import messages
from django.contrib.auth.models import User
from django.utils.translation import gettext as _
from .models import UserProfile, Student, Teacher, Course, Group, Attendance, Notification, Subject, Schedule, LeaveRequest
from .forms import StudentRegistrationForm, NotificationForm, LeaveRequestForm, UserProfileForm, UserUpdateForm, PasswordChangeCustomForm
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from datetime import datetime, date, timedelta
from dal import autocomplete
from django import forms
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

# ============= NOTIFICATION HELPER FUNCTIONS =============

def create_notification(recipient, notification_type, title, message, sender=None, student=None, leave_request=None):
    """Билдирме түзүү үчүн helper функция"""
    return Notification.objects.create(
        recipient=recipient,
        sender=sender,
        notification_type=notification_type,
        title=title,
        message=message,
        student=student,
        leave_request=leave_request
    )

def check_excessive_absences():
    """Көп жолу келбеген студенттер үчүн билдирмелер жөнөтүү"""
    from collections import defaultdict
    
    # Соңку 10 күндө 3 жолудан көп келбеген студенттерди тапуу
    last_10_days = date.today() - timedelta(days=10)
    
    student_absences = defaultdict(int)
    recent_absences = Attendance.objects.filter(
        date__gte=last_10_days,
        status='Absent'
    )
    
    for absence in recent_absences:
        student_absences[absence.student] += 1
    
    # 3тен көп жок болгон студенттер үчүн билдирме
    for student, absence_count in student_absences.items():
        if absence_count >= 3:
            # Ошол студенттин ата-энелерине билдирме жөнөтүү
            parent_profiles = student.parents.all()
            for parent_profile in parent_profiles:
                create_notification(
                    recipient=parent_profile.user,
                    notification_type='ABSENCE',
                    title=f'⚠️ {student.name} көп жолу келбеди',
                    message=f'Сиздин балаңыз {student.name} соңку 10 күн ичинде {absence_count} жолу сабактан келген жок. Анын себебин аныктап, зарыл чараларды көрүүгө өтүнөбүз.',
                    student=student
                )
            
            # Класс жетекчисине да билдирме
            if student.group:
                # Студенттин группасы менен иштеген мугалимдерди тапуу
                group_teachers = Teacher.objects.filter(
                    subject__schedule__group=student.group
                ).distinct()
                
                for teacher in group_teachers:
                    if teacher.user:
                        create_notification(
                            recipient=teacher.user,
                            notification_type='ABSENCE',
                            title=f'⚠️ {student.name} көп жолу келбеди',
                            message=f'Студент {student.name} ({student.group.name}) соңку 10 күн ичинде {absence_count} жолу сабактан келген жок.',
                            student=student
                        )

def send_leave_request_notification(leave_request):
    """Бошотуу сурамы жөнөтүлгөндө билдирме"""
    # Администраторлорго жана менеджерлерге билдирме
    admins_and_managers = User.objects.filter(
        userprofile__role__in=['ADMIN', 'MANAGER']
    )
    
    for admin in admins_and_managers:
        create_notification(
            recipient=admin,
            notification_type='LEAVE_REQUEST',
            title=f'📝 Жаңы бошотуу сурамы - {leave_request.student.name}',
            message=f'Студент {leave_request.student.name} ({leave_request.student.group.name}) {leave_request.start_date} - {leave_request.end_date} мөөнөтүндө {leave_request.get_leave_type_display()} себебинен бошотуу сурап жатат.\n\nСебеби: {leave_request.reason}',
            sender=leave_request.student.user,
            student=leave_request.student,
            leave_request=leave_request
        )

def send_leave_decision_notification(leave_request, approved_by):
    """Бошотуу сурамы боюнча чечим кабыл алынгандан кийин билдирме"""
    status_text = 'бекитилди' if leave_request.status == 'APPROVED' else 'четке кагылды'
    emoji = '✅' if leave_request.status == 'APPROVED' else '❌'
    
    # Студентке билдирме
    create_notification(
        recipient=leave_request.student.user,
        notification_type='LEAVE_APPROVED' if leave_request.status == 'APPROVED' else 'LEAVE_REJECTED',
        title=f'{emoji} Бошотуу сурамыңыз {status_text}',
        message=f'Сиздин {leave_request.start_date} - {leave_request.end_date} мөөнөтүндөгү бошотуу сурамыңыз {status_text}.\n\nЧечим кабыл алган: {approved_by.get_full_name() or approved_by.username}',
        sender=approved_by,
        student=leave_request.student,
        leave_request=leave_request
    )
    
    # Ата-энелерге да билдирме
    parent_profiles = leave_request.student.parents.all()
    for parent_profile in parent_profiles:
        create_notification(
            recipient=parent_profile.user,
            notification_type='LEAVE_APPROVED' if leave_request.status == 'APPROVED' else 'LEAVE_REJECTED',
            title=f'{emoji} {leave_request.student.name}дын бошотуу сурамы {status_text}',
            message=f'Балаңыздын {leave_request.start_date} - {leave_request.end_date} мөөнөтүндөгү бошотуу сурамы {status_text}.\n\nЧечим кабыл алган: {approved_by.get_full_name() or approved_by.username}',
            sender=approved_by,
            student=leave_request.student,
            leave_request=leave_request
        )

@login_required
def mark_schedule_attendance(request, group_id, period):
    if request.user.userprofile.role != 'TEACHER':
        return JsonResponse({'error': 'Бул функция мугалимдер үчүн гана.'}, status=403)
    
    group = get_object_or_404(Group, id=group_id)
    students = Student.objects.filter(group=group)
    today = date.today()
    
    data = {
        'group_name': group.name,
        'students': [
            {
                'id': student.id,
                'name': student.name,
                'attendance': Attendance.objects.filter(
                    student=student, schedule__group=group, schedule__day=today.weekday(), date=today
                ).exists()
            } for student in students
        ]
    }
    return JsonResponse(data)

@csrf_exempt
@login_required
def submit_schedule_attendance(request):
    if request.user.userprofile.role != 'TEACHER':
        return JsonResponse({'error': 'Бул функция мугалимдер үчүн гана.'}, status=403)
    
    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('status_'):
                student_id = key.replace('status_', '')
                student = get_object_or_404(Student, id=student_id)
                schedule_id = request.POST.get('schedule_id')
                schedule = get_object_or_404(Schedule, id=schedule_id)
                Attendance.objects.update_or_create(
                    student=student,
                    schedule=schedule,
                    date=date.today(),
                    defaults={'status': value or 'Absent', 'created_by': request.user}
                )
            elif key.startswith('late_'):
                student_id = key.replace('late_', '')
                student = get_object_or_404(Student, id=student_id)
                schedule_id = request.POST.get('schedule_id')
                schedule = get_object_or_404(Schedule, id=schedule_id)
                if value:
                    Attendance.objects.update_or_create(
                        student=student,
                        schedule=schedule,
                        date=date.today(),
                        defaults={'status': value, 'created_by': request.user}
                    )
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'Invalid request'}, status=400)
# Форма
class NotificationForm(forms.ModelForm):
    student = forms.ModelChoiceField(queryset=Student.objects.all(), label='Студент')
    message = forms.CharField(widget=forms.Textarea, label='Билдирүү')

    class Meta:
        model = Notification
        fields = ['student', 'message']

# Autocomplete
class UserProfileAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return UserProfile.objects.none()
        qs = UserProfile.objects.filter(role='PARENT')
        if self.q:
            qs = qs.filter(user__username__istartswith=self.q)
        return qs

# Ролдорду текшерүү
def is_admin_or_manager(user):
    try:
        return user.userprofile.role in ['ADMIN', 'MANAGER']
    except:
        return False

def is_teacher(user):
    try:
        return user.userprofile.role == 'TEACHER'
    except:
        return False

# View'дор
def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')

@csrf_exempt
def register(request):
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            UserProfile.objects.create(user=user, role='STUDENT')
            student = Student.objects.create(
                user=user,
                name=form.cleaned_data['name'],
                course=form.cleaned_data['course'],
                group=form.cleaned_data['group']
            )
            parent_username = form.cleaned_data.get('parent_username')
            if parent_username:
                parent = User.objects.get(username=parent_username)
                student.parents.add(parent.userprofile)
                parent.userprofile.students.add(student)
                messages.success(request, _('Parent linked successfully!'))
            messages.success(request, _('Student registered successfully!'))
            login(request, user)
            return redirect('dashboard')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = StudentRegistrationForm()
    return render(request, 'registration.html', {'form': form})

from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt
def user_login(request):
    if request.method == 'POST':
        # CSRF debug маалыматы
        print(f"POST request from user: {request.user}")
        print(f"CSRF cookie: {request.META.get('CSRF_COOKIE')}")
        print(f"POST CSRF token: {request.POST.get('csrfmiddlewaretoken')}")
        
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, _('Login error!'))
    return render(request, 'login.html')

def user_logout(request):
    auth_logout(request)
    messages.success(request, _('You have been logged out!'))
    return redirect('login')

@login_required
def settings(request):
    """
    User settings бети - тил жана тема тандоо
    """
    from django.conf import settings as django_settings
    
    if request.method == 'POST':
        # Тилди өзгөртүү
        if 'language' in request.POST:
            language = request.POST.get('language')
            if language in [lang[0] for lang in django_settings.LANGUAGES]:
                from django.utils import translation
                translation.activate(language)
                request.session['django_language'] = language
                messages.success(request, _('Language updated successfully!'))
        
        # Теманы өзгөртүү (JavaScript аркылуу иштейт)
        if 'theme' in request.POST:
            theme = request.POST.get('theme')
            if theme in ['light', 'dark']:
                messages.success(request, _('Theme preference saved!'))
        
        # Башка настройкалар
        if 'notifications' in request.POST:
            notifications_enabled = request.POST.get('notifications') == 'on'
            # Notifications настройкаларын сактоо логикасын бул жерге кошсо болот
            messages.success(request, _('Notification settings updated!'))
        
        return redirect('settings')
    
    # Жеткиликтүү тилдер
    available_languages = getattr(django_settings, 'LANGUAGES', [])
    current_language = request.session.get('django_language', django_settings.LANGUAGE_CODE)
    
    context = {
        'available_languages': available_languages,
        'current_language': current_language,
        'user': request.user,
        'user_profile': getattr(request.user, 'userprofile', None),
    }
    
    return render(request, 'settings.html', context)

@login_required
@login_required
def dashboard(request):
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    print(f"Dashboard view called for user: {request.user}")
    
    try:
        profile, created = UserProfile.objects.get_or_create(user=request.user, defaults={'role': 'STUDENT'})
        if created:
            messages.info(request, _('Your profile has been automatically created (STUDENT role).'))
        role = profile.role
        print(f"User role: {role}")
    except Exception as e:
        print(f"Error getting user profile: {e}")
        return HttpResponse(f"Error: {e}", status=500)
    
    # Бүгүнкү дата
    today = timezone.now().date()
    today_weekday_num = today.weekday()  # 0=Monday, 6=Sunday
    
    # Күндөрдү Django Schedule моделинин форматына алмаштыруу
    weekday_map = {
        0: 'MONDAY',
        1: 'TUESDAY', 
        2: 'WEDNESDAY',
        3: 'THURSDAY',
        4: 'FRIDAY',
        5: 'SATURDAY',
        6: 'SUNDAY'
    }
    today_weekday = weekday_map.get(today_weekday_num, 'MONDAY')
    
    context = {
        'role': role, 
        'today': today,
        'user': request.user
    }
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if role in ['ADMIN', 'MANAGER']:
        # Статистика маалыматтары
        total_students = Student.objects.count()
        total_teachers = Teacher.objects.count()
        total_groups = Group.objects.count()
        total_subjects = Subject.objects.count()
        
        # Соңку иш-аракеттер
        recent_activities = []
        
        context.update({
            'total_students': total_students,
            'total_teachers': total_teachers,
            'total_groups': total_groups,
            'total_subjects': total_subjects,
            'recent_activities': recent_activities,
        })
        
    elif role == 'TEACHER':
        # Мугалим профилин алуу
        try:
            teacher = Teacher.objects.get(user=request.user)
            context['teacher_obj'] = teacher
            
            # Мугалимдин предметтери
            my_subjects = Subject.objects.filter(teacher=teacher)
            my_subjects_count = my_subjects.count()
            
            # Мугалимдин студенттери (группалар аркылуу)
            my_groups = Group.objects.filter(schedule__subject__teacher=teacher).distinct()
            my_students = Student.objects.filter(group__in=my_groups)
            my_students_count = my_students.count()
            
            # Бүгүнкү расписание
            today_schedule = Schedule.objects.filter(
                subject__teacher=teacher,
                day=today_weekday
            ).order_by('time_slot__order')
            
            today_classes_count = today_schedule.count()
            
            context.update({
                'my_subjects_count': my_subjects_count,
                'my_students_count': my_students_count,
                'today_classes_count': today_classes_count,
                'today_schedule': today_schedule,
            })
            
        except Teacher.DoesNotExist:
            messages.error(request, _('Teacher profile not found. Please contact administrator.'))
            
    elif role == 'STUDENT':
        # Студент профилин алуу же түзүү
        student, created = Student.objects.get_or_create(
            user=request.user,
            defaults={
                'name': request.user.get_full_name() or request.user.username,
                'course': Course.objects.first(),
                'group': Group.objects.first()
            }
        )
        
        if created and not Course.objects.exists():
            Course.objects.create(name='1-курс', year=2024)
        if created and not Group.objects.exists():
            course = Course.objects.first()
            Group.objects.create(name='А-группа', course=course)
            
        context['student_obj'] = student
        
        # Катышуу статистикасы
        attendance = Attendance.objects.filter(student=student)
        if start_date:
            attendance = attendance.filter(date__gte=start_date)
        if end_date:
            attendance = attendance.filter(date__lte=end_date)
            
        total_attendance = attendance.count()
        present_days = attendance.filter(status='Present').count()
        absent_days = attendance.filter(status='Absent').count()
        excused_days = attendance.filter(status='Excused').count()
        
        attendance_percentage = (present_days / total_attendance * 100) if total_attendance > 0 else 0
        
        # Бүгүнкү расписание
        student_today_schedule = []
        if student.group:
            student_today_schedule = Schedule.objects.filter(
                group=student.group,
                day=today_weekday
            ).order_by('time_slot__order')
            
            # Ар бир сабак үчүн катышуу статусун алуу
            for schedule in student_today_schedule:
                try:
                    attendance_record = Attendance.objects.get(
                        student=student,
                        date=today,
                        subject=schedule.subject
                    )
                    schedule.attendance_status = attendance_record.status
                except Attendance.DoesNotExist:
                    schedule.attendance_status = None
        
        context.update({
            'attendance_percentage': round(attendance_percentage, 1),
            'present_days': present_days,
            'absent_days': absent_days,
            'excused_days': excused_days,
            'student_today_schedule': student_today_schedule,
            'start_date': start_date,
            'end_date': end_date
        })
        
    elif role == 'PARENT':
        # Ата-энелердин балдары
        my_children = profile.students.all()
        
        if my_children:
            # Ар бир бала үчүн статистика
            for child in my_children:
                attendance = Attendance.objects.filter(student=child)
                if start_date:
                    attendance = attendance.filter(date__gte=start_date)
                if end_date:
                    attendance = attendance.filter(date__lte=end_date)
                    
                total_attendance = attendance.count()
                present_count = attendance.filter(status='Present').count()
                absent_count = attendance.filter(status='Absent').count()
                
                child.present_count = present_count
                child.absent_count = absent_count
                child.attendance_percentage = (present_count / total_attendance * 100) if total_attendance > 0 else 0
            
            # Ата-энелер үчүн билдирүүлөр
            parent_notifications = Notification.objects.filter(
                recipient=request.user,
                notification_type__in=['ABSENCE', 'LEAVE_REQUEST', 'GENERAL']
            ).order_by('-created_at')[:10]
            
            context.update({
                'my_children': my_children,
                'parent_notifications': parent_notifications,
                'start_date': start_date,
                'end_date': end_date
            })
        else:
            messages.error(request, _('Your child profile is not linked. Please contact administrator.'))
            context.update({'message': 'Баланын маалыматы жок'})

    print(f"Dashboard context keys: {list(context.keys())}")
    print(f"Rendering dashboard.html template")
    
    # Context маалыматтарын толук чыгаралы  
    print(f"Student object: {context.get('student_obj')}")
    if 'student_obj' in context and context['student_obj']:
        student = context['student_obj']
        print(f"Student details: name={student.name}, group={student.group}")
        if hasattr(student, 'group') and student.group:
            print(f"Group details: {student.group.name}")
    
    try:
        response = render(request, 'dashboard.html', context)
        print(f"Template rendered successfully, content length: {len(response.content)}")
        return response
    except Exception as e:
        print(f"Template rendering error: {e}")
        import traceback
        print(f"Full traceback: {traceback.format_exc()}")
        return HttpResponse(f"Template error: {e}", status=500)

@login_required
def send_notification(request):
    if request.user.userprofile.role != 'TEACHER':
        messages.error(request, _('This function is for teachers only.'))
        return redirect('dashboard')
    
    teacher = Teacher.objects.get(user=request.user)
    if request.method == 'POST':
        form = NotificationForm(request.POST)
        if form.is_valid():
            notification = form.save(commit=False)
            notification.teacher = teacher
            notification.save()
            messages.success(request, _('Notification sent!'))
            return redirect('dashboard')
    else:
        form = NotificationForm()
    return render(request, 'send_notification.html', {'form': form})

@login_required
def mark_notification_read(request, notification_id):
    if request.user.userprofile.role != 'PARENT':
        messages.error(request, _('This function is for parents only.'))
        return redirect('dashboard')
    notification = Notification.objects.get(id=notification_id)
    if notification.student in request.user.userprofile.students.all():
        notification.is_read = True
        notification.save()
        messages.success(request, _('Notification marked as read.'))
    return redirect('dashboard')

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import Schedule, Subject, Group
from .forms import ScheduleEditForm  # Форма файлын импорттоо

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Schedule, Subject, Group
from .forms import ScheduleEditForm  # Жаңы кошулган

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import Course, Group, Schedule, Teacher, Attendance, Student, Subject
from datetime import date
import json

@login_required
@login_required
def schedule_edit(request):
    """Расписание көрүү (Менеджер/Админ үчүн - жөн гана көрүү режими)"""
    profile = UserProfile.objects.get(user=request.user)
    
    # Уруксат текшерүү - жөн гана көрүү үчүн
    if profile.role not in ['MANAGER', 'ADMIN']:
        messages.error(request, _('This function is for managers and administrators only.'))
        return redirect('dashboard')

    schedules = Schedule.objects.all()  # Бардык расписаниелер
    courses = Course.objects.all()
    
    # Убакыт периоддору (статик)
    periods = []
    times = [
        ('08:00', '09:30'),
        ('09:40', '11:10'), 
        ('11:30', '13:00'),
        ('14:00', '15:30'),
        ('15:40', '17:10')
    ]
    
    for i, (start, end) in enumerate(times, 1):
        from datetime import datetime
        start_time = datetime.strptime(start, '%H:%M').time()
        end_time = datetime.strptime(end, '%H:%M').time()
        periods.append({
            'period': i,
            'start_time': start_time,
            'end_time': end_time
        })

    # Көрүү режиминде POST өндөшүү жок - жөн гана маалымат көрсөтүү
    context = {
        'courses': courses,
        'active_course': courses.first() if courses.exists() else None,
        'periods': periods,
        'schedules': schedules,
        'view_only': True,  # Көрүү режими белгиси
    }
    return render(request, 'schedule_edit.html', context)

@login_required
def schedule_update(request):
    if request.method == 'POST' and request.user.userprofile.role == 'MANAGER':
        try:
            data = json.loads(request.body)
            group_id = data.get('group_id')
            period = data.get('period')
            content = data.get('content')

            # Расписаниени жаңылоо логикасы
            # Мисалы, content'ти бөлүп, subject, teacher жана room'ду сактоо
            lines = content.strip().split('\n')
            subject_name = lines[0] if lines else ''
            teacher_name = lines[1] if len(lines) > 1 else ''
            room = lines[2] if len(lines) > 2 else ''

            group = get_object_or_404(Group, id=group_id)
            subject = Subject.objects.filter(subject_name=subject_name).first()
            teacher = Teacher.objects.filter(name=teacher_name).first()

            Schedule.objects.update_or_create(
                group=group,
                period=period,
                defaults={
                    'subject': subject,
                    'teacher': teacher,
                    'room': room,
                    'start_time': '10:00',  # Убакытты моделден же periods тизмесинен алса болот
                    'end_time': '11:20'
                }
            )
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
def student_schedule(request):
    """Студенттин жумалык расписаниеси жана катышуу статистикасы"""
    if request.user.userprofile.role != 'STUDENT':
        messages.error(request, _('This function is for students only.'))
        return redirect('dashboard')
    
    try:
        student = Student.objects.get(user=request.user)
        
        # Жумалык расписание үчүн күндөр
        days = [
            ('Monday', 'Дүйшөмбү'),
            ('Tuesday', 'Шейшемби'), 
            ('Wednesday', 'Шаршемби'),
            ('Thursday', 'Бейшемби'),
            ('Friday', 'Жума'),
            ('Saturday', 'Ишемби'),
        ]
        
        # Студенттин группасынын расписаниеси
        schedules = Schedule.objects.filter(group=student.group).order_by('day', 'time_slot__order')
        
        # Расписание матрицасын түзүү
        schedule_matrix = {}
        
        # Ушул аптанын башы жана аягы (Дүйшөмбүдөн Жекшембиге)
        from datetime import date, timedelta
        today = date.today()
        # Дүйшөмбүнү табуу (weekday: 0=Дүйшөмбү, 6=Жекшемби)
        days_since_monday = today.weekday()
        start_date = today - timedelta(days=days_since_monday)
        end_date = start_date + timedelta(days=6)  # Жекшемби
        
        # TimeSlot'торду алуу (админ тарабынан конфигурацияланган)
        from core.models import TimeSlot
        time_slots = TimeSlot.objects.filter(is_active=True).order_by('order')
        
        for day_code, day_name in days:
            schedule_matrix[day_code] = {}
            for time_slot in time_slots:
                # Бул күн жана убакыт слоту үчүн сабакты табуу
                day_schedule = schedules.filter(
                    day=day_code,
                    time_slot=time_slot
                ).first()
                
                schedule_info = None
                if day_schedule:
                    # Бул сабак үчүн ушул аптанын эң соңку катышуу статусун алуу
                    last_attendance = Attendance.objects.filter(
                        student=student,
                        subject=day_schedule.subject,
                        date__range=[start_date, end_date]
                    ).order_by('-date').first()
                    
                    # Соңку статусту аныктоо
                    if last_attendance:
                        if last_attendance.status == 'Present':
                            last_status = 'Катышкан'
                            status_class = 'success'
                        elif last_attendance.status == 'Absent':
                            last_status = 'Катышпаган'
                            status_class = 'danger'
                        elif last_attendance.status == 'Late':
                            last_status = 'Кечиккен'
                            status_class = 'warning'
                        else:
                            last_status = 'Катышкан'
                            status_class = 'success'
                    else:
                        last_status = 'Белгилене элек'
                        status_class = 'secondary'
                    
                    schedule_info = {
                        'schedule': day_schedule,
                        'last_status': last_status,
                        'status_class': status_class
                    }
                
                schedule_matrix[day_code][time_slot.name] = {
                    'time_range': f"{time_slot.start_time.strftime('%H:%M')} - {time_slot.end_time.strftime('%H:%M')}",
                    'schedule_info': schedule_info
                }
        
        # Бүгүнкү расписание үчүн schedules алуу
        today_schedules = Schedule.objects.filter(
            group=student.group,
            day=today.strftime('%A')  # Monday, Tuesday, etc.
        ).order_by('time_slot__order')
        
        context = {
            'student': student,
            'schedules': today_schedules,
            'days': days,
            'time_slots': list(time_slots),  # TimeSlot QuerySet'ин list'ке айландыруу
            'schedule_matrix': schedule_matrix,
            'start_date': start_date,
            'end_date': end_date,
        }
        
        return render(request, 'modern_student_schedule.html', context)
        
    except Student.DoesNotExist:
        messages.error(request, _('Student information not found!'))
        return redirect('dashboard')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Course, Group, Schedule, Teacher, Attendance, Student
from datetime import date

@login_required
def teacher_schedule(request):
    """Мугалимдин расписаниесин көрүү жана жоктоо белгилөө"""
    if request.user.userprofile.role != 'TEACHER':
        messages.error(request, _('This function is for teachers only.'))
        return redirect('dashboard')
    
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, _('You are not registered as a teacher.'))
        return redirect('dashboard')

    # Фильтрлер
    day_filter = request.GET.get('day_filter', '')
    group_filter = request.GET.get('group_filter', '')
    subject_filter = request.GET.get('subject_filter', '')

    # Мугалимдин сабактары
    schedules = Schedule.objects.filter(
        subject__teacher=teacher
    ).select_related('subject', 'group', 'group__course').order_by('day', 'time_slot__order')
    
    # Бүгүнкү сабактар үчүн фильтр
    today = date.today()
    today_schedules = schedules.filter(day=today.strftime('%A'))

    # Фильтрлерди колдонуу
    if day_filter:
        today_schedules = today_schedules.filter(day=day_filter)
    if group_filter:
        today_schedules = today_schedules.filter(group_id=group_filter)
    if subject_filter:
        today_schedules = today_schedules.filter(subject_id=subject_filter)

    # Статистика үчүн маалыматтар
    total_lessons = schedules.count()
    available_groups = Group.objects.filter(schedule__subject__teacher=teacher).distinct()
    available_subjects = teacher.subjects.all()
    total_groups = available_groups.count()
    total_subjects = available_subjects.count()

    # Күндөр тизмеси
    days = [
        ('Monday', 'Дүйшөмбү'),
        ('Tuesday', 'Шейшемби'), 
        ('Wednesday', 'Шаршемби'),
        ('Thursday', 'Бейшемби'),
        ('Friday', 'Жума'),
        ('Saturday', 'Ишемби'),
    ]

    context = {
        'teacher': teacher,
        'schedules': today_schedules,
        'total_lessons': total_lessons,
        'total_groups': total_groups,
        'total_subjects': total_subjects,
        'available_groups': available_groups,
        'available_subjects': available_subjects,
        'days': days,
        'day_filter': day_filter,
        'group_filter': group_filter,
        'subject_filter': subject_filter,
    }
    return render(request, 'modern_teacher_schedule.html', context)

@login_required
def teacher_attendance(request):
    """Мугалим үчүн жеке жоктоо системасы"""
    if request.user.userprofile.role != 'TEACHER':
        messages.error(request, _('This function is for teachers only.'))
        return redirect('dashboard')
    
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, _('You are not registered as a teacher.'))
        return redirect('dashboard')

    # POST request - жоктоо белгилөө
    if request.method == 'POST':
        schedule_id = request.POST.get('schedule_id')
        if schedule_id:
            schedule = get_object_or_404(Schedule, id=schedule_id, subject__teacher=teacher)
            
            saved_count = 0
            for key, value in request.POST.items():
                if key.startswith('status_') and value:
                    try:
                        student_id = key.replace('status_', '')
                        student = get_object_or_404(Student, id=student_id, group=schedule.group)
                        
                        # Attendance сактоо
                        attendance, created = Attendance.objects.update_or_create(
                            student=student,
                            subject=schedule.subject,
                            schedule=schedule,
                            date=date.today(),
                            defaults={
                                'status': value, 
                                'created_by': request.user
                            }
                        )
                        saved_count += 1
                        
                    except (Student.DoesNotExist, ValueError):
                        continue
            
            if saved_count > 0:
                messages.success(request, _('Attendance marked for {} students.').format(saved_count))
            else:
                messages.warning(request, 'Эч кандай маалымат сакталган жок.')
                
            return redirect('teacher_attendance')

    # GET request - мугалимдин сабактарын көрсөтүү
    schedules = Schedule.objects.filter(
        subject__teacher=teacher
    ).select_related(
        'subject', 'group', 'group__course'
    ).prefetch_related(
        'group__student_set'
    ).order_by('day', 'time_slot__order')
    
    # Статистика эсептөө
    total_students = sum([schedule.group.student_set.count() for schedule in schedules])
    unique_groups = schedules.values('group').distinct().count()
    subjects = schedules.values('subject').distinct()

    context = {
        'teacher': teacher,
        'schedules': schedules,
        'total_students': total_students,
        'unique_groups': unique_groups,
        'subjects': subjects,
        'today': date.today(),
    }
    return render(request, 'teacher_attendance.html', context)
@login_required
def mark_attendance(request, subject_id):
    if request.user.userprofile.role not in ['TEACHER', 'ADMIN', 'MANAGER']:
        messages.error(request, _('This function is for teachers or administrators only.'))
        return redirect('dashboard')
    
    subject = get_object_or_404(Subject, id=subject_id)
    students = Student.objects.filter(group=subject.course.group_set.first())
    
    if request.method == 'POST':
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status:
                Attendance.objects.create(
                    student=student,
                    subject=subject,
                    date=date.today(),
                    status=status,
                    created_by=request.user
                )
        messages.success(request, 'Катышуу белгиленди.')
        return redirect('schedule')  # Бул жерде багыттоо туура эмес, анткени 'schedule' жок. 'dashboard' же 'teacher_schedule' колдонуңуз.
    
    context = {'subject': subject, 'students': students}
    return render(request, 'mark_attendance.html', context)

@login_required
def mark_group_attendance(request, schedule_id):
    if request.user.userprofile.role != 'TEACHER':
        messages.error(request, 'Бул функция мугалимдер үчүн гана.')
        return redirect('dashboard')
    
    schedule = get_object_or_404(Schedule, id=schedule_id)
    students = Student.objects.filter(group=schedule.group)
    
    if request.method == 'POST':
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status:
                Attendance.objects.update_or_create(
                    student=student,
                    subject=schedule.subject,  # Subject кошуу
                    schedule=schedule,
                    date=date.today(),
                    defaults={'status': status, 'created_by': request.user}
                )
        messages.success(request, 'Катышуу белгиленди.')
        return redirect('teacher_schedule')
    
    context = {'schedule': schedule, 'students': students}
    return render(request, 'mark_group_attendance.html', context)

@login_required
@user_passes_test(is_admin_or_manager)
def report(request):
    from django.utils import timezone
    from datetime import datetime, timedelta
    import json
    
    # Дата фильтри
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    report_type = request.GET.get('report_type', 'all')  # all, daily, weekly, monthly
    student_id = request.GET.get('student')
    group_id = request.GET.get('group')
    
    # Базалык attendance queryset
    attendances = Attendance.objects.all()
    
    # Дата боюнча фильтр
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            attendances = attendances.filter(date__gte=start_date)
        except ValueError:
            start_date = None
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            attendances = attendances.filter(date__lte=end_date)
        except ValueError:
            end_date = None
    
    # Студент боюнча фильтр
    if student_id:
        try:
            attendances = attendances.filter(student_id=student_id)
        except:
            pass
    
    # Группа боюнча фильтр
    if group_id:
        try:
            attendances = attendances.filter(student__group_id=group_id)
        except:
            pass
    
    # Быстрые фильтры
    today = timezone.now().date()
    if report_type == 'daily':
        attendances = attendances.filter(date=today)
    elif report_type == 'weekly':
        week_start = today - timedelta(days=today.weekday())
        attendances = attendances.filter(date__gte=week_start, date__lte=today)
    elif report_type == 'monthly':
        month_start = today.replace(day=1)
        attendances = attendances.filter(date__gte=month_start, date__lte=today)
    
    # Статистика
    total_records = attendances.count()
    present_count = attendances.filter(status='Present').count()
    absent_count = attendances.filter(status='Absent').count()
    late_count = attendances.filter(status='Late').count()
    excused_count = attendances.filter(status='Excused').count()
    
    attendance_percentage = (present_count / total_records * 100) if total_records > 0 else 0
    
    # Группа боюнча статистика
    group_stats = []
    for group in Group.objects.all():
        group_attendances = attendances.filter(student__group=group)
        group_total = group_attendances.count()
        group_present = group_attendances.filter(status='Present').count()
        group_percentage = (group_present / group_total * 100) if group_total > 0 else 0
        
        group_stats.append({
            'group': group,
            'total': group_total,
            'present': group_present,
            'percentage': round(group_percentage, 1)
        })
    
    # Күнүмдүк тренд (соңку 7 күн)
    daily_trend = []
    for i in range(7):
        check_date = today - timedelta(days=i)
        day_attendances = Attendance.objects.filter(date=check_date)
        day_total = day_attendances.count()
        day_present = day_attendances.filter(status='Present').count()
        day_percentage = (day_present / day_total * 100) if day_total > 0 else 0
        
        daily_trend.append({
            'date': check_date.strftime('%m-%d'),
            'percentage': round(day_percentage, 1),
            'total': day_total
        })
    
    daily_trend.reverse()  # Хронологиялык тартипте
    
    # Эң көп жок болгон студенттер
    from django.db.models import Count
    top_absent_students = Student.objects.annotate(
        absent_count=Count('attendance', filter=Q(attendance__status='Absent'))
    ).filter(absent_count__gt=0).order_by('-absent_count')[:10]
    
    # Export handling
    if request.method == 'POST':
        format_type = request.POST.get('format')
        if format_type == 'pdf':
            return export_detailed_pdf(attendances, {
                'report_type': report_type,
                'start_date': start_date,
                'end_date': end_date,
                'total_records': total_records,
                'present_count': present_count,
                'absent_count': absent_count,
                'late_count': late_count,
                'excused_count': excused_count,
                'attendance_percentage': attendance_percentage
            })
        elif format_type == 'excel':
            return export_detailed_excel(attendances)
    
    context = {
        'attendances': attendances.order_by('-date')[:100],  # Show latest 100
        'total_records': total_records,
        'present_count': present_count,
        'absent_count': absent_count, 
        'late_count': late_count,
        'excused_count': excused_count,
        'attendance_percentage': round(attendance_percentage, 1),
        'group_stats': group_stats,
        'daily_trend': json.dumps(daily_trend),
        'top_absent_students': top_absent_students,
        'students': Student.objects.all(),
        'groups': Group.objects.all(),
        'start_date': start_date,
        'end_date': end_date,
        'report_type': report_type,
        'student_id': student_id,
        'group_id': group_id,
    }
    
    return render(request, 'report.html', context)

def export_detailed_pdf(attendances, stats):
    """Детальный PDF отчет с статистикой"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="detailed_attendance_report.pdf"'
    
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfutils
    from reportlab.lib.units import inch
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph("Катышуу Системасы - Детальный Отчет", title_style))
    elements.append(Spacer(1, 20))
    
    # Statistics Summary
    summary_data = [
        ['Жалпы жазуулар:', stats['total_records']],
        ['Катышкандар:', stats['present_count']],
        ['Катышпагандар:', stats['absent_count']], 
        ['Кечиккендер:', stats['late_count']],
        ['Уруксат менен жоктор:', stats['excused_count']],
        ['Катышуу пайызы:', f"{stats['attendance_percentage']}%"],
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Attendance Records Table
    if attendances.exists():
        elements.append(Paragraph("Катышуу Жазуулары", styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        data = [['Студент', 'Сабак', 'Дата', 'Статус']]
        for att in attendances[:50]:  # Limit to 50 records for PDF
            data.append([
                att.student.name,
                att.subject.subject_name if att.subject else 'Н/Д',
                att.date.strftime('%d.%m.%Y'),
                att.get_status_display()
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
    doc.build(elements)
    return response

def export_detailed_excel(attendances):
    """Детальный Excel отчет"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="detailed_attendance_report.xlsx"'
    
    workbook = Workbook()
    
    # Main data sheet
    worksheet = workbook.active
    worksheet.title = "Катышуу Жазуулары"
    
    # Headers
    headers = ['Студент', 'Группа', 'Сабак', 'Дата', 'Статус', 'Жаратылган күнү', 'Жараткан']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Data
    for row_num, att in enumerate(attendances, 2):
        worksheet.cell(row=row_num, column=1, value=att.student.name)
        worksheet.cell(row=row_num, column=2, value=att.student.group.name if att.student.group else 'Н/Д')
        worksheet.cell(row=row_num, column=3, value=att.subject.subject_name if att.subject else 'Н/Д')
        worksheet.cell(row=row_num, column=4, value=att.date.strftime('%d.%m.%Y'))
        worksheet.cell(row=row_num, column=5, value=att.get_status_display())
        worksheet.cell(row=row_num, column=6, value=att.date.strftime('%d.%m.%Y'))
        worksheet.cell(row=row_num, column=7, value=att.created_by.username if att.created_by else 'Система')
    
    # Statistics sheet
    stats_sheet = workbook.create_sheet("Статистика")
    stats_data = [
        ['Көрсөткүч', 'Мааниси'],
        ['Жалпы жазуулар', attendances.count()],
        ['Катышкандар', attendances.filter(status='Present').count()],
        ['Катышпагандар', attendances.filter(status='Absent').count()],
        ['Кечиккендер', attendances.filter(status='Late').count()],
        ['Уруксат менен жоктор', attendances.filter(status='Excused').count()],
    ]
    
    for row_num, (label, value) in enumerate(stats_data, 1):
        stats_sheet.cell(row=row_num, column=1, value=label)
        stats_sheet.cell(row=row_num, column=2, value=value)
    
    workbook.save(response)
    return response

def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="report.pdf"'
    p = canvas.Canvas(response)
    p.drawString(100, 750, "Катышуу Отчету")
    y = 700
    for att in Attendance.objects.all()[:20]:
        p.drawString(100, y, f"{att.student.name} - {att.status} - {att.date}")
        y -= 20
    p.showPage()
    p.save()
    return response

def export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Студент'
    ws['B1'] = 'Сабак'
    ws['C1'] = 'Статус'
    ws['D1'] = 'Дата'
    row = 2
    for att in Attendance.objects.all():
        ws[f'A{row}'] = att.student.name
        ws[f'B{row}'] = att.subject.subject_name if att.subject else 'Н/Д'
        ws[f'C{row}'] = att.status
        ws[f'D{row}'] = str(att.date)
        row += 1
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
    wb.save(response)
    return response

# Жалпы расписание view (ролдорго жараша багыттоо)
@login_required
def schedule(request):
    """
    Расписание главном бети - жаңы unified системага багытталат
    """
    # Бардык ролдор үчүн жаңы unified schedule системасына багыттоо
    return redirect('unified_schedule')

# ============= БОШОТУУ СУРАМДАРЫ =============

@login_required
def submit_leave_request(request):
    """Студент бошотуу сурамын жөнөтөт"""
    if request.user.userprofile.role != 'STUDENT':
        messages.error(request, 'Бул функция студенттер үчүн гана.')
        return redirect('dashboard')
    
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, 'Студент профили табылган жок.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.student = student
            leave_request.save()
            
            # Автоматтык билдирме жөнөтүү
            send_leave_request_notification(leave_request)
            
            messages.success(request, 'Бошотуу сурамыңыз ийгиликтүү жөнөтүлдү. Администрацияга билдирме жөнөтүлдү.')
            return redirect('my_leave_requests')
    else:
        form = LeaveRequestForm()
    
    return render(request, 'leave/submit_request.html', {'form': form})

@login_required
def my_leave_requests(request):
    """Студенттин өз бошотуу сурамдарын көрүү"""
    if request.user.userprofile.role != 'STUDENT':
        messages.error(request, 'Бул функция студенттер үчүн гана.')
        return redirect('dashboard')
    
    try:
        student = Student.objects.get(user=request.user)
        leave_requests = LeaveRequest.objects.filter(student=student).order_by('-created_at')
    except Student.DoesNotExist:
        leave_requests = []
        messages.error(request, 'Студент профили табылган жок.')
    
    return render(request, 'leave/my_requests.html', {'leave_requests': leave_requests})

@login_required
@user_passes_test(lambda u: u.userprofile.role in ['ADMIN', 'MANAGER', 'TEACHER'])
def manage_leave_requests(request):
    """Администраторлор/менеджерлер бошотуу сурамдарын башкарат"""
    pending_requests = LeaveRequest.objects.filter(status='PENDING').order_by('-created_at')
    all_requests = LeaveRequest.objects.all().order_by('-created_at')
    
    context = {
        'pending_requests': pending_requests,
        'all_requests': all_requests,
    }
    return render(request, 'leave/manage_requests.html', context)

@login_required
@user_passes_test(lambda u: u.userprofile.role in ['ADMIN', 'MANAGER', 'TEACHER'])
def approve_leave_request(request, request_id):
    """Бошотуу сурамын бекитүү"""
    leave_request = get_object_or_404(LeaveRequest, id=request_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            leave_request.status = 'APPROVED'
            leave_request.approved_by = request.user
            leave_request.save()
            
            # Бекитилген мөөнөтүндө автоматтык "Excused" белгиси коюу
            current_date = leave_request.start_date
            while current_date <= leave_request.end_date:
                # Ошол студенттин ошол күндөгү сабактары үчүн Excused деп белгилөө
                schedules = Schedule.objects.filter(group=leave_request.student.group)
                for schedule in schedules:
                    Attendance.objects.update_or_create(
                        student=leave_request.student,
                        subject=schedule.subject,
                        date=current_date,
                        defaults={
                            'status': 'Excused',
                            'created_by': request.user,
                            'leave_request': leave_request
                        }
                    )
                current_date += timedelta(days=1)
            
            # Билдирме жөнөтүү
            send_leave_decision_notification(leave_request, request.user)
            
            messages.success(request, f'{leave_request.student.name}дын бошотуу сурамы бекитилди. Студентке жана ата-энелерине билдирме жөнөтүлдү.')
        
        elif action == 'reject':
            leave_request.status = 'REJECTED' 
            leave_request.approved_by = request.user
            leave_request.save()
            
            # Билдирме жөнөтүү
            send_leave_decision_notification(leave_request, request.user)
            
            messages.info(request, f'{leave_request.student.name}дын бошотуу сурамы четке кагылды. Студентке жана ата-энелерине билдирме жөнөтүлдү.')
    
    return redirect('manage_leave_requests')

# ============= NOTIFICATION VIEWS =============

@login_required
def notifications(request):
    """Колдонуучунун билдирмелерин көрсөтүү"""
    user_notifications = Notification.objects.filter(recipient=request.user)
    unread_count = user_notifications.filter(is_read=False).count()
    
    return render(request, 'notifications/list.html', {
        'notifications': user_notifications,
        'unread_count': unread_count
    })

@login_required
def mark_notification_read(request, notification_id):
    """Билдирмени окулган деп белгилөө"""
    notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    notification.is_read = True
    notification.save()
    
    if request.headers.get('HX-Request'):
        # HTMX request болсо
        return JsonResponse({'success': True})
    
    return redirect('notifications')

@login_required
def mark_all_notifications_read(request):
    """Бардык билдирмелерди окулган деп белгилөө"""
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    messages.success(request, 'Бардык билдирмелер окулган деп белгиленди.')
    return redirect('notifications')

@login_required
@user_passes_test(lambda u: u.userprofile.role in ['ADMIN', 'MANAGER'])
def send_bulk_notification(request):
    """Көптөгөн колдонуучуларга билдирме жөнөтүү"""
    if request.method == 'POST':
        title = request.POST.get('title')
        message = request.POST.get('message')
        recipient_type = request.POST.get('recipient_type')  # all_students, all_parents, specific_group
        group_id = request.POST.get('group_id')
        
        recipients = []
        
        if recipient_type == 'all_students':
            recipients = User.objects.filter(userprofile__role='STUDENT')
        elif recipient_type == 'all_parents':
            recipients = User.objects.filter(userprofile__role='PARENT')
        elif recipient_type == 'specific_group' and group_id:
            recipients = User.objects.filter(
                student__group_id=group_id
            )
        
        # Билдирмелерди түзүү
        notifications_to_create = []
        for recipient in recipients:
            notifications_to_create.append(
                Notification(
                    recipient=recipient,
                    sender=request.user,
                    notification_type='GENERAL',
                    title=title,
                    message=message
                )
            )
        
        # Bulk create для эффективности
        Notification.objects.bulk_create(notifications_to_create)
        
        messages.success(request, f'{len(recipients)} колдонуучуга билдирме жөнөтүлдү.')
        return redirect('notifications')
    
    groups = Group.objects.all()
    return render(request, 'notifications/bulk_send.html', {'groups': groups})

# ============= SCHEDULE MANAGEMENT =============

@login_required
@login_required
def manage_schedule(request):
    """Расписание башкаруу (Админ/Менеджер үчүн)"""
    profile = UserProfile.objects.get(user=request.user)
    
    # Уруксат текшерүү
    if profile.role not in ['ADMIN', 'MANAGER']:
        messages.error(request, 'Сизге бул бетке кирүүгө уруксат жок.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            subject_id = request.POST.get('subject_id')
            group_id = request.POST.get('group_id')
            day = request.POST.get('day')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            
            if not all([subject_id, group_id, day, start_time, end_time]):
                return JsonResponse({'error': 'Бардык талапка сай маалыматты толтуруңуз'}, status=400)
            
            subject = Subject.objects.get(id=subject_id)
            group = Group.objects.get(id=group_id)
            
            # Убакыт кайталануусун текшерүү
            conflict = Schedule.objects.filter(
                group=group,
                day=day,
                start_time__lt=end_time,
                end_time__gt=start_time
            ).exists()
            
            if conflict:
                return JsonResponse({'error': 'Бул убакытта группада башка сабак бар'}, status=400)
            
            # Мугалимдин убакыт кайталануусун текшерүү
            teacher_conflict = Schedule.objects.filter(
                subject__teacher=subject.teacher,
                day=day,
                start_time__lt=end_time,
                end_time__gt=start_time
            ).exists()
            
            if teacher_conflict:
                return JsonResponse({'error': 'Бул убакытта мугалим башка группада сабак өтөт'}, status=400)
            
            # Жаңы расписание түзүү
            schedule = Schedule.objects.create(
                subject=subject,
                group=group,
                day=day,
                start_time=start_time,
                end_time=end_time
            )
            
            return JsonResponse({'success': True, 'message': 'Сабак ийгиликтүү кошулду'})
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    # Филтрлер
    course_filter = request.GET.get('course')
    group_filter = request.GET.get('group') 
    teacher_filter = request.GET.get('teacher')
    day_filter = request.GET.get('day')
    
    # Бардык маалыматтар
    courses = Course.objects.all().order_by('year')
    groups = Group.objects.select_related('course').all().order_by('course__year', 'name')
    teachers = Teacher.objects.all().order_by('name')
    subjects = Subject.objects.select_related('teacher', 'course').all()
    
    # Курс боюнча иерархия түзүү
    courses_hierarchy = []
    
    for course in courses:
        course_groups = groups.filter(course=course)
        
        if course_filter and str(course.id) != course_filter:
            continue
            
        groups_data = []
        for group in course_groups:
            if group_filter and str(group.id) != group_filter:
                continue
                
            # Группанын расписаниеси
            schedules = Schedule.objects.filter(group=group).select_related(
                'subject__teacher'
            ).order_by('day', 'time_slot__order')
            
            # Филтрлерди колдонуу
            if teacher_filter:
                schedules = schedules.filter(subject__teacher_id=teacher_filter)
            if day_filter:
                schedules = schedules.filter(day=day_filter)
            
            # Студенттердин саны
            students_count = Student.objects.filter(group=group).count()
            
            groups_data.append({
                'group': group,
                'schedules': schedules,
                'students_count': students_count
            })
        
        if groups_data:  # Эгер группалар бар болсо гана кошуу
            courses_hierarchy.append({
                'course': course,
                'groups': groups_data
            })
    
    context = {
        'courses_hierarchy': courses_hierarchy,
        'courses': courses,
        'groups': groups,
        'all_groups': groups,  # Modal үчүн
        'teachers': teachers,
        'subjects': subjects,
    }
    
    return render(request, 'manage_schedule.html', context)

@login_required 
def delete_schedule(request, schedule_id):
    """Расписание өчүрүү"""
    profile = UserProfile.objects.get(user=request.user)
    
    if profile.role not in ['ADMIN', 'MANAGER']:
        return JsonResponse({'error': 'Уруксат жок'}, status=403)
    
    try:
        schedule = Schedule.objects.get(id=schedule_id)
        schedule.delete()
        return JsonResponse({'success': True, 'message': 'Сабак ийгиликтүү өчүрүлдү'})
    except Schedule.DoesNotExist:
        return JsonResponse({'error': 'Сабак табылган жок'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def manage_subjects(request):
    """Сабактар (Предметтер) көрүү панели"""
    profile = UserProfile.objects.get(user=request.user)
    
    # Уруксат текшерүү - бардык роль кире алат, бирок ADMIN/MANAGER гана өзгөртө алат
    if profile.role not in ['ADMIN', 'MANAGER', 'TEACHER']:
        messages.error(request, 'Сизге бул бетке кирүүгө уруксат жок.')
        return redirect('dashboard')
    
    # Филтрлер
    course_filter = request.GET.get('course')
    teacher_filter = request.GET.get('teacher')
    search = request.GET.get('search', '').strip()
    
    # Негизги маалыматтар
    subjects = Subject.objects.select_related('teacher', 'course').all()
    courses = Course.objects.all().order_by('year')
    teachers = Teacher.objects.all().order_by('name')
    
    # Филтрлерди колдонуу
    if course_filter:
        subjects = subjects.filter(course_id=course_filter)
    if teacher_filter:
        subjects = subjects.filter(teacher_id=teacher_filter)
    if search:
        subjects = subjects.filter(subject_name__icontains=search)
    
    subjects = subjects.order_by('course__year', 'subject_name')
    
    # Статистика эсептөө
    total_subjects = subjects.count()
    subjects_by_course = {}
    subjects_by_teacher = {}
    
    for subject in subjects:
        # Курс боюнча топтоо
        course_name = subject.course.name
        if course_name not in subjects_by_course:
            subjects_by_course[course_name] = []
        subjects_by_course[course_name].append(subject)
        
        # Мугалим боюнча топтоо
        teacher_name = subject.teacher.name if subject.teacher else "Мугалим дайындалган жок"
        if teacher_name not in subjects_by_teacher:
            subjects_by_teacher[teacher_name] = []
        subjects_by_teacher[teacher_name].append(subject)
    
    # Ар бир сабак үчүн кошумча маалымат
    subjects_data = []
    for subject in subjects:
        # Бул сабак боюнча расписание саны
        schedule_count = Schedule.objects.filter(subject=subject).count()
        
        # Бул сабак боюнча студенттердин саны
        students_count = Student.objects.filter(
            group__schedule__subject=subject
        ).distinct().count()
        
        subjects_data.append({
            'subject': subject,
            'schedule_count': schedule_count,
            'students_count': students_count
        })
    
    context = {
        'subjects_data': subjects_data,
        'courses': courses,
        'teachers': teachers,
        'total_subjects': total_subjects,
        'subjects_by_course': subjects_by_course,
        'subjects_by_teacher': subjects_by_teacher,
        'course_filter': course_filter,
        'teacher_filter': teacher_filter,
        'search': search,
        'can_edit': profile.role in ['ADMIN', 'MANAGER']
    }
    
    return render(request, 'manage_subjects.html', context)

@login_required
@login_required
@login_required
def universal_schedule(request):
    """Универсалдуу расписание көрүү - бардык ролдор үчүн (жаңыланган версия)"""
    try:
        profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user, role='STUDENT')
    
    # Бардык расписанияны алуу (негизги query)
    schedules_query = Schedule.objects.select_related(
        'subject__teacher',
        'group__course'
    ).prefetch_related(
        'subject',
        'group'
    ).all()
    
    # Тартипке салуу - күн, убакыт, группа боюнча
    schedules = schedules_query.order_by('day', 'time_slot__order', 'group__name')    # Күндөр тизмеси тартипке салынган
    days_choices = [
        ('Monday', 'Дүйшөмбү'),
        ('Tuesday', 'Шейшемби'),
        ('Wednesday', 'Шаршемби'), 
        ('Thursday', 'Бейшемби'),
        ('Friday', 'Жума'),
        ('Saturday', 'Ишемби'),
        ('Sunday', 'Жекшемби')
    ]    # Күн тартиби боюнча сортировка (1=Дүйшөмбү, 7=Жекшемби)
    day_order = {
        'Monday': 1,
        'Tuesday': 2,
        'Wednesday': 3,
        'Thursday': 4,
        'Friday': 5,
        'Saturday': 6,
        'Sunday': 7
    }
    
    # Тартипке салуу - алгач күн, анан убакыт, анан группа
    schedules = schedules_query.extra(
        select={
            'day_order': f"CASE day "
                        f"WHEN 'Monday' THEN 1 "
                        f"WHEN 'Tuesday' THEN 2 "
                        f"WHEN 'Wednesday' THEN 3 "
                        f"WHEN 'Thursday' THEN 4 "
                        f"WHEN 'Friday' THEN 5 "
                        f"WHEN 'Saturday' THEN 6 "
                        f"WHEN 'Sunday' THEN 7 "
                        f"END"
        }
    ).order_by('day_order', 'time_slot__order', 'group__name')
    
    # Күн боюнча группалап чыгаруу
    schedules_by_day = {}
    for day_code, day_name in days_choices:
        day_schedules = schedules.filter(day=day_code)
        if day_schedules.exists():
            schedules_by_day[day_code] = {
                'name': day_name,
                'schedules': day_schedules,
                'count': day_schedules.count()
            }
    
    # Статистика
    total_schedules = schedules.count()
    unique_groups = schedules.values('group').distinct().count()
    unique_teachers = schedules.values('subject__teacher').distinct().count()
    unique_subjects = schedules.values('subject').distinct().count()
    
    # Бүгүнкү расписание
    from datetime import date
    today = date.today()
    today_day_name = today.strftime('%A')  # Monday, Tuesday, etc.
    today_schedules = schedules.filter(day=today_day_name)
    
    context = {
        # Негизги маалыматтар (жаңы template үчүн)
        'schedules': schedules,
        'schedules_by_day': schedules_by_day,  # Күн боюнча группаланган
        'today_schedules': today_schedules,
        'today_day_name': today_day_name,
        
        # Статистика
        'total_schedules': total_schedules,
        'unique_groups': unique_groups,
        'total_teachers': unique_teachers,
        'total_subjects': unique_subjects,
        'total_rooms': len(schedules.values_list('room', flat=True).distinct()),
        
        # Колдонуучу маалыматы
        'user_role': profile.role,
        'today': today,
    }
    
    return render(request, 'universal_schedule.html', context)

def weekly_schedule(request):
    """Жумалык расписание таблицасы - ар күн үчүн так сабактар"""
    profile = UserProfile.objects.get(user=request.user)
    
    # Бардык роль үчүн ачык, бирок ар бир роль өз маалыматын көрөт
    
    # Күндөр тизмеси - студенттер үчүн жума күндөрү гана
    if profile.role == 'STUDENT':
        days = [
            ('Monday', 'Дүйшөмбү'),
            ('Tuesday', 'Шейшемби'), 
            ('Wednesday', 'Шаршемби'),
            ('Thursday', 'Бейшемби'),
            ('Friday', 'Жума'),
        ]
    else:
        # Башка роллор үчүн бардык күндөр
        days = [
            ('Monday', 'Дүйшөмбү'),
            ('Tuesday', 'Шейшемби'), 
            ('Wednesday', 'Шаршемби'),
            ('Thursday', 'Бейшемби'),
            ('Friday', 'Жума'),
            ('Saturday', 'Ишемби'),
        ]
    
    # Роль боюнча филтер жана жума күндөрү
    schedules_query = Schedule.objects.select_related('subject__teacher', 'group__course').all()
    
    if profile.role == 'STUDENT':
        try:
            student = Student.objects.get(user=request.user)
            # Студенттер үчүн жума күндөрү гана
            weekdays = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
            schedules_query = schedules_query.filter(
                group=student.group,
                day__in=weekdays
            )
        except Student.DoesNotExist:
            schedules_query = Schedule.objects.none()
    elif profile.role == 'TEACHER':
        try:
            teacher = Teacher.objects.get(user=request.user)
            schedules_query = schedules_query.filter(subject__teacher=teacher)
        except Teacher.DoesNotExist:
            schedules_query = Schedule.objects.none()
    # MANAGER/ADMIN үчүн бардык расписаниелер көрүнөт
    
    # Филтрлер (MANAGER/ADMIN үчүн)
    course_filter = request.GET.get('course')
    group_filter = request.GET.get('group')
    teacher_filter = request.GET.get('teacher')
    
    if course_filter:
        schedules_query = schedules_query.filter(group__course_id=course_filter)
    if group_filter:
        schedules_query = schedules_query.filter(group_id=group_filter)
    if teacher_filter:
        schedules_query = schedules_query.filter(subject__teacher_id=teacher_filter)
    
    schedules = schedules_query.order_by('time_slot__order')
    
    # Таблица үчүн маалыматтарды уюштуруу
    schedule_matrix = {}
    
    # TimeSlot'торду алуу (админ тарабынан конфигурацияланган)
    from core.models import TimeSlot
    time_slots = TimeSlot.objects.filter(is_active=True).order_by('order')
    
    for day_code, day_name in days:
        schedule_matrix[day_code] = {}
        for time_slot in time_slots:
            # Бул күн жана убакыт слоту үчүн сабактарды табуу
            day_schedules = schedules.filter(
                day=day_code,
                time_slot=time_slot
            )
            
            schedule_matrix[day_code][time_slot.name] = {
                'time_range': f"{time_slot.start_time.strftime('%H:%M')} - {time_slot.end_time.strftime('%H:%M')}",
                'schedules': day_schedules
            }
    
    # Филтр үчүн маалыматтар
    courses = Course.objects.all().order_by('year') if profile.role in ['ADMIN', 'MANAGER'] else []
    groups = Group.objects.select_related('course').all().order_by('course__year', 'name') if profile.role in ['ADMIN', 'MANAGER'] else []
    teachers = Teacher.objects.all().order_by('name') if profile.role in ['ADMIN', 'MANAGER'] else []
    
    context = {
        'days': days,
        'time_slots': list(time_slots),  # TimeSlot QuerySet'ин list'ке айландыруу
        'schedule_matrix': schedule_matrix,
        'courses': courses,
        'groups': groups,
        'teachers': teachers,
        'course_filter': course_filter,
        'group_filter': group_filter,
        'teacher_filter': teacher_filter,
        'user_role': profile.role,
    }
    
    return render(request, 'weekly_schedule.html', context)


# ============= ПРОФИЛ ЖАНА КОЛДОНУУЧУ БАШКАРУУ =============

@login_required
def profile_view(request):
    """Колдонуучунун профилин көрсөтүү"""
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        # Эгер профил жок болсо, жаңысын түзүү
        profile = UserProfile.objects.create(user=request.user)
    
    context = {
        'profile': profile,
        'user': request.user,
    }
    return render(request, 'profile/profile_view.html', context)


@login_required
def profile_edit(request):
    """Профилди өзгөртүү"""
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = UserProfileForm(request.POST, request.FILES, instance=profile)
        
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            # Профилдин толуктугун текшерүү
            profile.check_profile_completeness()
            messages.success(request, 'Профилиңиз ийгиликтүү жаңыртылды!')
            return redirect('profile_view')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = UserProfileForm(instance=profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'profile': profile,
    }
    return render(request, 'profile/profile_edit.html', context)


@login_required
def change_password(request):
    """Сыр сөздү өзгөртүү"""
    if request.method == 'POST':
        form = PasswordChangeCustomForm(user=request.user, data=request.POST)
        if form.is_valid():
            request.user.set_password(form.cleaned_data['new_password1'])
            request.user.save()
            messages.success(request, 'Сыр сөзүңүз ийгиликтүү өзгөртүлдү! Кайрадан кириңиз.')
            return redirect('login')
    else:
        form = PasswordChangeCustomForm(user=request.user)

    context = {
        'form': form,
    }
    return render(request, 'profile/change_password.html', context)


@login_required
def profile_photo_delete(request):
    """Профил сүрөтүн өчүрүү"""
    if request.method == 'POST':
        try:
            profile = request.user.userprofile
            if profile.profile_photo:
                profile.profile_photo.delete()
                profile.save()
                messages.success(request, 'Профил сүрөтү ийгиликтүү өчүрүлдү.')
            else:
                messages.info(request, 'Профил сүрөтү жок.')
        except UserProfile.DoesNotExist:
            messages.error(request, 'Профил табылган жок.')
        
        return redirect('profile_edit')
    
    return redirect('profile_view')


@login_required  
def profile_completion_check(request):
    """Профилдин толуктугун текшерүү (AJAX)"""
    try:
        profile = request.user.userprofile
        is_complete = profile.check_profile_completeness()
        
        completion_data = {
            'is_complete': is_complete,
            'missing_fields': [],
            'completion_percentage': 0
        }
        
        # Керектүү талаалар
        required_fields = [
            ('first_name', 'Аты', request.user.first_name),
            ('last_name', 'Фамилиясы', request.user.last_name),
            ('email', 'Email', request.user.email),
            ('phone_number', 'Телефон', profile.phone_number),
        ]
        
        filled_count = 0
        total_count = len(required_fields)
        
        for field_name, field_label, field_value in required_fields:
            if not field_value:
                completion_data['missing_fields'].append(field_label)
            else:
                filled_count += 1
        
        completion_data['completion_percentage'] = int((filled_count / total_count) * 100)
        
        return JsonResponse(completion_data)
        
    except UserProfile.DoesNotExist:
        return JsonResponse({
            'is_complete': False,
            'missing_fields': ['Профил табылган жок'],
            'completion_percentage': 0
        })


def get_user_profile_or_create(user):
    """Колдонуучунун профилин алуу же жок болсо жаңысын түзүү"""
    try:
        return user.userprofile
    except UserProfile.DoesNotExist:
        return UserProfile.objects.create(user=user)


@login_required
@csrf_exempt
def mark_student_attendance(request, schedule_id):
    """Студент өзүнүн катышуусун белгилөө"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        schedule = get_object_or_404(Schedule, id=schedule_id)
        student_profile = request.user.userprofile
        
        # Студент экенин текшерүү
        if student_profile.role != 'STUDENT':
            return JsonResponse({'error': 'Only students can mark their own attendance'}, status=403)
        
        # Студенттин бул группада экенин текшерүү
        try:
            student = Student.objects.get(user=request.user)
            if student.group != schedule.group:
                return JsonResponse({'error': 'You are not enrolled in this class'}, status=403)
        except Student.DoesNotExist:
            return JsonResponse({'error': 'Student profile not found'}, status=404)
        
        # Учурдагу жума үчүн гана катышуу белгилөөгө уруксат берүү
        from datetime import datetime, timedelta
        today = datetime.now().date()
        
        # Жуманын башталышы (дүйшөмбү)
        days_since_monday = today.weekday()
        week_start = today - timedelta(days=days_since_monday)
        week_end = week_start + timedelta(days=6)
        
        # Schedule күнүн алуу
        schedule_date = today  # Азырынча бүгүнкү күн үчүн
        
        if not (week_start <= schedule_date <= week_end):
            return JsonResponse({'error': 'Attendance can only be marked for current week'}, status=400)
        
        # JSON маалыматтарды алуу
        data = json.loads(request.body)
        status = data.get('status')
        
        if status not in ['Present', 'Absent', 'Late']:
            return JsonResponse({'error': 'Invalid attendance status'}, status=400)
        
        # Attendance жазууну түзүү же жаңылоо
        attendance, created = Attendance.objects.update_or_create(
            student=student,
            schedule=schedule,
            date=schedule_date,
            defaults={
                'status': status,
                'marked_by_student': True,
                'marked_at': datetime.now()
            }
        )
        
        return JsonResponse({
            'success': True,
            'status': status,
            'message': f'Катышуу "{status}" деп белгиленди',
            'created': created
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error marking attendance: {str(e)}'}, status=500)